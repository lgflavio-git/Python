import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import Reference


wb = xl.load_workbook(filename='transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(row=1, column=4)
cell.value = "corrected_price_longtexttomake big"



for n in range(2, sheet.max_row + 1):
    cell = sheet.cell(row=n, column=3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row=n, column=4)
    corrected_price_cell.value = corrected_price



for col in sheet.columns:
    print(col[0].column_letter)
    sheet.column_dimensions[col[0].column_letter].auto_size = True
    #sheet.column_dimensions[col[0].column_letter].width = 100

#sheet.column_dimensions["D"].bestFit = True
#sheet.column_dimensions["D"].width = 100


# values = Reference(sheet=sheet,
#                   min_row=1,
#                   max_row=sheet.max_row,
#                   min_column=1,
#                   max_columns=sheet.max_column,
#                   ))

# Access a range of cells
cell_range = sheet["A1:D4"]
for column in cell_range:
    for cell in column:
        print(cell.value)

wb.save('transactions2.xlsx')
